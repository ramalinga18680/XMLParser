from lxml import etree
import sys
import openpyxl


class Analyzer():
    def __init__(self):
        self.test_case = ""
        self.ts_case_obj = None
        self.report_ws = None
        self.row_to_write = 2
    '''

    def writerow(self, data, row_to_write):
        i = 0
        max_range: int = 1 + len(data)
        for col in range(1, max_range):
            self.report_ws.cell(row_to_write, col).value = data[i]
            i += 1
    '''
    def writerow(self, data,row_to_write=None):
        self.report_ws.append(data)

    def write_header(self):

        if self.report_ws is not None:
            data = ["Test Pack", "Test Case Id", "Test Result", "Total Steps", "Total Steps Passed",
                    "Total Steps Failed", "Has Ignition", "Ignition Status", "Has Audio Verification",
                    "Audio Verification Status"]
            self.writerow(data, 1)

    def initialize(self, report):
        self.report_ws = report
        self.write_header()

    def analyze(self, test_case):
        print("Analyzing the test case")
        self.ts_case_obj = TestCase(test_case, self)
        self.ts_case_obj.process_test_steps()
        self.ts_case_obj.process_func_units()
        self.generate_report()

        # list_func_units = ts_case.xpath('./functionUnit[functionUnitName[contains(text(),"Ignition")]]')
        del self.ts_case_obj
        self.ts_case_obj = None

    def generate_report(self):
        data = []
        if self.report_ws is not None:
            data.append(self.ts_case_obj.find_test_pack_details())
            data.append(self.ts_case_obj.find_test_caseid())
            data.append(self.ts_case_obj.find_test_case_result())
            data.append(self.ts_case_obj.num_total_steps)
            data.append(self.ts_case_obj.num_passed_test_steps)
            data.append(self.ts_case_obj.num_failed_test_steps)

            if self.ts_case_obj.has_ignition:
                data.append("Yes")
                if self.ts_case_obj.ignition_failure is True:
                    data.append("Fail")
                else:
                    data.append("Pass")
            else:
                data.append("No")
                data.append("NA")

            if self.ts_case_obj.has_audio_verification:
                data.append("Yes")
                if self.ts_case_obj.verify_audio_failure is True:
                    data.append("Fail")
                else:
                    data.append("Pass")
            else:
                data.append("No")
                data.append("NA")

            #self.writerow(data, self.row_to_write)
            self.writerow(data)
            self.row_to_write += 1
        print("test_pack : tc_id : tc_req_id : tc_result", self.ts_case_obj.find_test_caseid(),
              self.ts_case_obj.find_test_case_reqid(), self.ts_case_obj.find_test_case_result(),
              self.ts_case_obj.find_test_pack_details())
        print(" Total steps :%d passed steps : %d Failed test steps :%d has ignition failure %d" % (
            self.ts_case_obj.num_total_steps, self.ts_case_obj.num_passed_test_steps,
            self.ts_case_obj.num_failed_test_steps, self.ts_case_obj.ignition_failure))
        if self.ts_case_obj.verify_audio_failure and self.ts_case_obj.has_audio_verification:
            print("Audio verification failed")

    def on_test_step_found(self, elem_ts_cmd, elem_ts_result):
        if "VERIFY_AUDIO" in elem_ts_cmd.text:
            self.ts_case_obj.has_audio_verification = True
            if elem_ts_result.text == "FAILURE":
                self.ts_case_obj.verify_audio_failure = True

        if elem_ts_result.text == "SUCCESS":
            self.ts_case_obj.num_passed_test_steps += 1
        elif elem_ts_result.text == "FAILURE":
            self.ts_case_obj.num_failed_test_steps += 1

    def on_func_unit_found(self, func_unit_name, func_unit_result):
        if "Ignition" in func_unit_name.text:
            self.ts_case_obj.has_ignition = True
            # print("Func unit name ", func_unit_name.text, func_unit_result.text)
            if func_unit_result.text == "FAILURE":
                self.ts_case_obj.ignition_failure = True


class TestCase():
    def __init__(self, ts_case_element, analyzer, out=sys.stdout):
        self.req_id = ""
        self.test_case_id = ""
        self.result = ""
        self.test_pack = ""
        self.ts_case_element = ts_case_element
        self.num_total_steps = 0
        self.num_passed_test_steps = 0
        self.num_failed_test_steps = 0
        self.has_ignition = False
        self.ignition_failure = False
        self.has_audio_verification = False
        self.verify_audio_failure = False
        self.analyzer = analyzer

    # def __del__(self):

    def find_test_caseid(self):
        elem_test_case_id = self.ts_case_element.find("./testCaseID")
        return elem_test_case_id.text

    def find_test_case_reqid(self):
        elem_test_case_req_id = self.ts_case_element.find("./requirementID")
        return elem_test_case_req_id.text

    def find_test_case_result(self):
        elem_test_case_result = self.ts_case_element.find("./result")
        return elem_test_case_result.text

    def find_test_pack_details(self):
        elem_test_pack_result = self.ts_case_element.find("./testPack")
        return elem_test_pack_result.text

    def find_test_steps(self):
        list_test_steps = self.ts_case_element.findall(".//TestStep")
        return list_test_steps

    def process_test_steps(self):
        list_steps = self.find_test_steps()
        self.num_total_steps = len(list_steps)
        for ts in list_steps:
            ts_result = ts.find("./result")
            ts_command = ts.find("./command")
            self.analyzer.on_test_step_found(ts_command, ts_result)

    def find_function_units(self):
        list_function_units = self.ts_case_element.findall("./functionUnit")
        # print("len of func units is ",len(list_func_units))
        return list_function_units

    def process_func_units(self):
        elements = ('functionUnitName', 'functionUnitResult')
        list_f_units = self.find_function_units()
        for f_unit in list_f_units:
            f_uname = f_unit.find(elements[0])
            f_result = f_unit.find(elements[1])
            self.analyzer.on_func_unit_found(f_uname, f_result)


if __name__ == '__main__':
    tree = etree.parse(sys.argv[1])

    report_wb = openpyxl.Workbook();
    report_sheet = report_wb.create_sheet(title='Analysis', index=0)

    analyzer_obj = Analyzer();
    analyzer_obj.initialize(report_sheet)

    list_test_case = tree.findall(".//TestCase")
    print("Num of test cases is", len(list_test_case))

    for ts_case in list_test_case:
        analyzer_obj.analyze(ts_case)

    report_wb.save(sys.argv[2])
    report_wb.close()
